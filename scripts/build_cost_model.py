"""Build docs/phase_c_cost_model.xlsx — a cost calculator grounded in the MEASURED
pilot token profile and current (2026) list prices. Formulas live in the sheet so
editing any price/assumption recomputes totals. Run: python scripts/build_cost_model.py
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

USD = '"$"#,##0.00'
INT = "#,##0"

# Measured from the pilot (cand_001 = 1 of 50 problems; full corpus = x50).
MEAS = dict(items=1900, gen_in=2263, gen_out=2262, vcr_turns_per_item=56 / 38,
            vcr_in=2000, vcr_out=1000, xcheck_in=2500, xcheck_out=800)

# Candidate eval models: (model, provider, $in/1M, $out/1M, reasoning?, confidence)
MODELS = [
    ("gpt-oss-120b", "Cerebras", 0.35, 0.75, "yes", "high"),
    ("qwen-3-235b-a22b", "Cerebras", 0.60, 1.20, "yes", "high"),
    ("llama-3.3-70b", "Cerebras", 0.85, 1.20, "no", "medium"),
    ("zai-glm-4.7", "Cerebras", 2.25, 2.75, "yes", "high"),
    ("gpt-4o-mini", "OpenAI", 0.15, 0.60, "no", "high"),
    ("gpt-4.1-mini", "OpenAI", 0.40, 1.60, "no", "high"),
    ("gpt-4.1", "OpenAI", 2.00, 8.00, "no", "high"),
    ("gpt-4o", "OpenAI", 2.50, 10.00, "no", "high"),
    ("claude-haiku-4.5", "Anthropic", 1.00, 5.00, "yes", "high"),
    ("claude-sonnet-4.6", "Anthropic", 3.00, 15.00, "yes", "high"),
    ("claude-opus-4.8", "Anthropic", 5.00, 25.00, "yes", "high"),
    ("gemini-2.5-flash", "Google", 0.30, 2.50, "yes", "high"),
    ("gemini-2.5-pro", "Google", 1.25, 10.00, "yes", "high"),
    ("deepseek-v3", "DeepSeek", 0.27, 1.10, "no", "medium"),
    ("deepseek-r1", "DeepSeek", 0.55, 2.19, "yes", "medium"),
]
# Candidate JUDGE models for the cheaper-judge scenarios
JUDGES = [("zai-glm-4.7 (used)", 2.25, 2.75), ("claude-haiku-4.5", 1.00, 5.00),
          ("gemini-2.5-flash", 0.30, 2.50), ("gpt-4o-mini", 0.15, 0.60)]

HEAD = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="305496")
BOLD = Font(bold=True)
TOTFILL = PatternFill("solid", fgColor="FCE4D6")

wb = openpyxl.Workbook()

# ---------------------------------------------------------------- Supuestos
s = wb.active
s.title = "Supuestos"
s["A1"] = "SycoCode — modelo de costes (base: tokens MEDIDOS en el piloto cand_001)"
s["A1"].font = Font(bold=True, size=13)
rows = [
    ("Parámetro", "Valor", "Fuente / nota"),
    ("Ítems del corpus completo", MEAS["items"], "50 problemas × 38 ítems"),
    ("Tokens prompt / ítem (gen)", MEAS["gen_in"], "MEDIDO (piloto)"),
    ("Tokens completion / ítem (gen)", MEAS["gen_out"], "MEDIDO (incluye razonamiento)"),
    ("VCR: nº de jueces", 3, "rúbrica (panel de 3)"),
    ("VCR: turnos juzgados / ítem", round(MEAS["vcr_turns_per_item"], 4), "MEDIDO (56/38)"),
    ("VCR: tokens IN / llamada", MEAS["vcr_in"], "ESTIMADO (editable)"),
    ("VCR: tokens OUT / llamada", MEAS["vcr_out"], "ESTIMADO (editable)"),
    ("xcheck: llamadas / ítem", 1, "MEDIDO"),
    ("xcheck: tokens IN / llamada", MEAS["xcheck_in"], "ESTIMADO (editable)"),
    ("xcheck: tokens OUT / llamada", MEAS["xcheck_out"], "ESTIMADO (editable)"),
    ("Juez: precio IN ($/1M)", 2.25, "zai-glm-4.7 (Cerebras)"),
    ("Juez: precio OUT ($/1M)", 2.75, "zai-glm-4.7 (Cerebras)"),
]
for i, (a, b, c) in enumerate(rows, start=3):
    s.cell(i, 1, a); s.cell(i, 2, b); s.cell(i, 3, c)
    if i == 3:
        for col in (1, 2, 3):
            s.cell(i, col).font = HEAD; s.cell(i, col).fill = HFILL
# named refs
ITEMS, GIN, GOUT = "$B$4", "$B$5", "$B$6"
NJ, VTPI, VIN, VOUT = "$B$7", "$B$8", "$B$9", "$B$10"
XPI, XIN, XOUT = "$B$11", "$B$12", "$B$13"
JIN, JOUT = "$B$14", "$B$15"
# derived block
s["A17"] = "DERIVADO"; s["A17"].font = BOLD
derived = [
    ("Llamadas VCR totales", f"={ITEMS}*{VTPI}*{NJ}", INT),
    ("Llamadas xcheck totales", f"={ITEMS}*{XPI}", INT),
    ("Coste VCR ($)", f"=B18*({VIN}*{JIN}+{VOUT}*{JOUT})/1000000", USD),
    ("Coste xcheck ($)", f"=B19*({XIN}*{JIN}+{XOUT}*{JOUT})/1000000", USD),
    ("COSTE JUEZ TOTAL ($) — fijo por modelo evaluado", "=B20+B21", USD),
    ("Coste generación gpt-oss (ref) ($)", f"={ITEMS}*({GIN}*0.35+{GOUT}*0.75)/1000000", USD),
]
for j, (a, f, fmt) in enumerate(derived, start=18):
    s.cell(j, 1, a); c = s.cell(j, 2, f); c.number_format = fmt
s["B22"].font = BOLD; s["B22"].fill = TOTFILL
s.column_dimensions["A"].width = 42; s.column_dimensions["B"].width = 14; s.column_dimensions["C"].width = 34
JUDGE_TOTAL = "Supuestos!$B$22"

# ---------------------------------------------------------- Coste por modelo
m = wb.create_sheet("Coste por modelo")
hdr = ["Modelo", "Proveedor", "$IN/1M", "$OUT/1M", "Razona", "Gen tokens IN",
       "Gen tokens OUT", "Coste GEN ($)", "Coste JUEZ ($)", "TOTAL 1 modelo ($)", "Conf."]
m.append(hdr)
for c in range(1, len(hdr) + 1):
    m.cell(1, c).font = HEAD; m.cell(1, c).fill = HFILL
for r, (name, prov, pin, pout, reas, conf) in enumerate(MODELS, start=2):
    m.cell(r, 1, name); m.cell(r, 2, prov); m.cell(r, 3, pin); m.cell(r, 4, pout); m.cell(r, 5, reas)
    m.cell(r, 6, f"=Supuestos!{ITEMS}*Supuestos!{GIN}").number_format = INT
    m.cell(r, 7, f"=Supuestos!{ITEMS}*Supuestos!{GOUT}").number_format = INT
    m.cell(r, 8, f"=(F{r}*C{r}+G{r}*D{r})/1000000").number_format = USD
    m.cell(r, 9, f"={JUDGE_TOTAL}").number_format = USD
    m.cell(r, 10, f"=H{r}+I{r}").number_format = USD
    m.cell(r, 10).font = BOLD; m.cell(r, 10).fill = TOTFILL
    m.cell(r, 11, conf)
m.append([])
note = m.max_row + 1
m.cell(note, 1, "TOTAL = generación (varía por modelo) + juez VCR/xcheck (fijo, usa zai-glm-4.7). "
                "Para N modelos evaluados: ~ N × (su generación + juez). El juez domina en modelos baratos.")
m.cell(note, 1).font = Font(italic=True)
widths = [20, 11, 9, 9, 8, 15, 15, 13, 13, 16, 7]
for i, w in enumerate(widths, start=1):
    m.column_dimensions[get_column_letter(i)].width = w

# ------------------------------------------------ Escenarios (recortar coste)
e = wb.create_sheet("Escenarios juez")
e["A1"] = "El juez VCR domina el coste. Palancas para recortarlo (coste JUEZ por modelo evaluado):"
e["A1"].font = BOLD
e.append([]); e.append(["Escenario", "Jueces", "Cobertura VCR", "Juez modelo", "$IN/1M", "$OUT/1M", "Coste JUEZ ($)"])
for c in range(1, 8):
    e.cell(3, c).font = HEAD; e.cell(3, c).fill = HFILL
# scenarios: vary judges, coverage fraction, judge model
scen = [
    ("Completo (actual)", 3, 1.0, JUDGES[0]),
    ("1 juez en vez de 3", 1, 1.0, JUDGES[0]),
    ("3 jueces, muestra 20%", 3, 0.2, JUDGES[0]),
    ("Juez barato (haiku-4.5), 3 jueces", 3, 1.0, JUDGES[1]),
    ("Juez barato (gemini-flash), 3 jueces", 3, 1.0, JUDGES[2]),
    ("Juez barato (gpt-4o-mini), 3 jueces", 3, 1.0, JUDGES[3]),
    ("Mínimo: 1 juez gpt-4o-mini, muestra 20%", 1, 0.2, JUDGES[3]),
]
for r, (lab, nj, cov, (jn, jin, jout)) in enumerate(scen, start=4):
    e.cell(r, 1, lab); e.cell(r, 2, nj); e.cell(r, 3, cov); e.cell(r, 4, jn)
    e.cell(r, 5, jin); e.cell(r, 6, jout)
    vcr = f"Supuestos!{ITEMS}*Supuestos!{VTPI}*{cov}*{nj}*(Supuestos!{VIN}*E{r}+Supuestos!{VOUT}*F{r})/1000000"
    xch = f"Supuestos!{ITEMS}*{cov}*(Supuestos!{XIN}*E{r}+Supuestos!{XOUT}*F{r})/1000000"
    e.cell(r, 7, f"={vcr}+{xch}").number_format = USD
    e.cell(r, 7).font = BOLD
for i, w in enumerate([38, 8, 14, 22, 9, 9, 14], start=1):
    e.column_dimensions[get_column_letter(i)].width = w

# ------------------------------------------------------------------- Notas
n = wb.create_sheet("Notas y fuentes")
notes = [
    "BASE: perfil de tokens MEDIDO en el piloto cand_001 (38 ítems, 68 requests de generación).",
    "  generación: 2263 tokens prompt + 2262 completion por ítem (incluye razonamiento de gpt-oss).",
    "  El corpus completo = 50 problemas = ×50 → ~8.6M tokens de generación, 8.400 llamadas VCR, 1.900 xcheck.",
    "Pass 2 (oráculo determinista) = GRATIS (ejecución local, sin API). No aparece en costes.",
    "Tokens IN/OUT por llamada del JUEZ son ESTIMADOS (editables en 'Supuestos'); el resto es medido.",
    "Los modelos de razonamiento (gpt-oss, glm-4.7, claude, gemini, qwen, r1) facturan los tokens de",
    "  razonamiento como OUTPUT → su coste real puede subir bastante vs un modelo no-razonador.",
    "Precios = lista pública 2026 (confianza alta salvo llama-3.3-70b y deepseek = media). VERIFICAR antes de presupuestar.",
    "Fuentes: cerebras.ai/pricing · platform.claude.com/docs/pricing · ai.google.dev/gemini-api/docs/pricing ·",
    "  pricepertoken.com (OpenAI) · api-docs.deepseek.com/quick_start/pricing.",
    "Cachés de prompt (OpenAI/Anthropic/DeepSeek/Gemini) y Batch API (-50%) NO están aplicados: reducirían el coste.",
    "DeepSeek 'chat'/'reasoner' se renombran a v4 el 2026-07-24 (más baratos); revisar si se usa DeepSeek.",
]
for i, t in enumerate(notes, start=1):
    n.cell(i, 1, t)
n.column_dimensions["A"].width = 110

import os
os.makedirs("docs", exist_ok=True)
wb.save("docs/phase_c_cost_model.xlsx")

# ---- also print a summary table (for the report), computed in Python (same formulas)
items, gin, gout = MEAS["items"], MEAS["gen_in"], MEAS["gen_out"]
gtin, gtout = items * gin, items * gout
jin, jout = 2.25, 2.75
vcr_calls = items * MEAS["vcr_turns_per_item"] * 3
xch_calls = items * 1
judge = (vcr_calls * (MEAS["vcr_in"] * jin + MEAS["vcr_out"] * jout) + xch_calls * (MEAS["xcheck_in"] * jin + MEAS["xcheck_out"] * jout)) / 1e6
print(f"JUDGE total (GLM-4.7, fixed/model): ${judge:.2f}  (VCR {vcr_calls:.0f} calls + xcheck {xch_calls:.0f})")
print(f"{'model':22s} {'gen$':>8s} {'+judge$':>8s} {'TOTAL$':>9s}")
for name, prov, pin, pout, reas, conf in MODELS:
    gen = (gtin * pin + gtout * pout) / 1e6
    print(f"{name:22s} {gen:8.2f} {judge:8.2f} {gen+judge:9.2f}")
print("\nCheap-judge scenarios (judge cost only):")
for jn, ji, jo in JUDGES:
    jc = (vcr_calls * (MEAS['vcr_in']*ji + MEAS['vcr_out']*jo) + xch_calls*(MEAS['xcheck_in']*ji + MEAS['xcheck_out']*jo))/1e6
    print(f"  3 judges, {jn:22s} ${jc:.2f}   | 1 judge ${jc/3:.2f}   | 1 judge+20% sample ${jc/3*0.2:.2f}")
print("\nsaved -> docs/phase_c_cost_model.xlsx")
